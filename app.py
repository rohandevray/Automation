
from os import execl
import openpyxl as xl
from openpyxl.chart import Reference, BarChart



def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    # cell = sheet.cell(1,1) its the way to point any cell in excel spreadsheet


    for row in range(2, sheet.max_row+1):
       cell = sheet.cell(row,3)
       corrected_value =(cell.value * 0.9)
       corrected_value_cell = sheet.cell(row,4)
       corrected_value_cell.value = corrected_value

    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()  # making a barchart structure named as chart
    chart.add_data(values) # adding the data to barchart
    sheet.add_chart(chart ,"e2") # e2 is (e,2) the coordinates we want to make a barchart 

    wb.save(filename) #we save all the new changes in the same workbook(wb) 